################################################################################
# Copyright (C) Troels Schwarz-Linnet - All Rights Reserved
# Written by Troels Schwarz-Linnet <tlinnet@gmail.com>, January 2018
# 
# Unauthorized copying of this file, via any medium is strictly prohibited.
#
# Any use of this code is strictly unauthorized without the written consent
# by the the author. This code is proprietary of the author.
# 
################################################################################

import datetime, os, os.path, tempfile, unittest
from openpyxl.utils import get_column_letter

# Import voxtools excel
from voxtools import excel


class Test_wb05(unittest.TestCase):
    def setUp(self):
        # Get the time
        self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W")
        # Create file obj
        file_obj = tempfile.NamedTemporaryFile(delete=True)
        self.excel_dst = file_obj.name + ".xlsx"
        # Get the folder with shared_data
        cur_dir = os.path.dirname(__file__)
        shared_data_dir = os.path.join(cur_dir, 'shared_data')
        # Get the filepath
        self.excel_src = os.path.join(shared_data_dir, 'wb05.xlsx')

        # Instantiate the Excel class
        self.exl = excel.excel(excel_src=self.excel_src, excel_dst=self.excel_dst)
        # Run it
        self.exl.run_all()


    def test_wb(self):
        # Assert the sheetnames
        self.assertEqual(self.exl.wb.sheetnames, ['Ark1', '(abs)', 'Køn'])


    def tearDown(self):
        # Delete temporary file
        print("Deleting : %s"%self.excel_dst)
        #os.remove(self.excel_dst)


if __name__ == '__main__':
    unittest.main()