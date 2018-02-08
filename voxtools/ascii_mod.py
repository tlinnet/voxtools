################################################################################
# Copyright (C) Voxmeter A/S - All Rights Reserved
#
# Voxmeter A/S
# Borgergade 6, 4.
# 1300 Copenhagen K
# Denmark
#
# Written by Troels Schwarz-Linnet <tsl@voxmeter.dk>, 2018
#
# Unauthorized copying of this file, via any medium is strictly prohibited.
#
# Any use of this code is strictly unauthorized without the written consent
# by Voxmeter A/S. This code is proprietary of Voxmeter A/S.
#
################################################################################
import shutil, datetime, copy, os, os.path, sys
from distutils.version import StrictVersion
import io, json

from openpyxl.utils import get_column_letter

# Check version
from openpyxl import __version__
test_version = StrictVersion(__version__) < StrictVersion("2.5.0")
if test_version:
    print("You need to have openpyxl version 2.5.0. You have %s"%__version__)
    sys.exit()


class create_ascii_input:
    def __init__(self, ascii_f=None, ascii_f_dst=None):
        # Store
        self.ascii_f = ascii_f
        self.ascii_f_dst = ascii_f_dst

        # Make current time
        #self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W-H%H-M%M-S%S")
        self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W")

    def run_all(self):
        # Copy file
        self.copy_file()

        # Read lines
        self.read_lines()

        # Find the questions
        self.find_sections()

        # Find questions
        self.find_Q()

        # Find type
        self.find_type()

        # Find pos
        self.find_pos()

        # Make header
        self.make_excel_header()

        # Make index
        self.make_index()

        # Clean dic
        self.clean_dic()
        # Save
        self.save_to_json()

    def make_index(self):
        # Loop over positions
        self.q_dic['index'] = {}
        for i, pos in enumerate(self.q_dic['pos']):
            question = self.q_dic['pos'][pos]
            # Get data
            q_data = self.q_dic[question]
            q_type = q_data['type']

            # If single
            if q_type in ["SQ", "NR"]:
                # Extract column letter
                q_col_let = q_data['col_let']
                q_col_text = ""
            # If multi
            elif q_type == "MQ":
                q_group_i = q_data['pos_group'][pos]
                q_col_let = q_data['col_let'][q_group_i]
                q_col_text = ""

            # Now run over the range
            pos_rep = pos.replace("(","").replace(")","")
            if "-" in pos_rep:
                pos_l, pos_h = pos_rep.split("-")
                pos_l = int(pos_l)
                pos_h = int(pos_h)
                j_range = list(range(pos_l, pos_h+1))
            else:
                j_range = [int(pos_rep)]

            # Loop over range
            for j in j_range:
                self.q_dic['index'][j-1] = (question, q_col_let)


    def save_to_json(self):
        # Save to json
        #with io.open(self.json_dst, 'w', encoding='utf-8') as f:
        with io.open(self.json_dst, 'w') as f:
            print("Saved to %s"%self.json_dst)
            #f.write(json.dumps(data, ensure_ascii=False))
            f.write(json.dumps(self.q_dic_clean, indent=2))

    def clean_dic(self):
        self.q_dic_clean = copy.copy(self.q_dic)
        for i, q in enumerate(self.questions):
            self.q_dic_clean[q].pop('section', None)

    def make_excel_header(self):
        # Loop over questions
        # Counter for letter
        if 'serial' in self.q_dic:
            col_i = 1
        else:
            col_i = 2

        for i, q in enumerate(self.questions):
            # Get the data
            q_data = self.q_dic[q]
            q_type = q_data['type']

            # If single
            if q_type in ["SQ", "NR"]:
                col_let = get_column_letter(col_i)
                self.q_dic[q]['col_let'] = col_let
                #print(q, col_let)
                col_i += 1

            # If multi
            elif q_type == "MQ":
                # There can be more letters
                col_letters = []
                q_groups = q_data['groups']
                for j in range(q_groups):
                    col_let = get_column_letter(col_i)
                    col_letters.append(col_let)
                    #print(q, col_let)
                    col_i += 1

                self.q_dic[q]['col_let'] = col_letters

    def find_pos(self):
        # Loop over questions
        for i, q in enumerate(self.questions):
            section = self.q_dic[q]['section']
            q_type = self.q_dic[q]['type']
            # Create storage
            self.q_dic[q]['pos'] = {}
            self.q_dic[q]['positions'] = []

            # For SQ
            if q_type == "SQ":
                cur_pos = None
                cur_pos_nr = []
                for j, line in enumerate(section):
                    line_s = line.split("|")
                    if len(line_s) > 1:
                        # Get the column
                        col = line_s[1].strip()
                        #if i == 0:
                        #    print(cur_pos, col, cur_pos_nr)

                        # If a position
                        if "(" in col:
                            # Store first
                            if cur_pos != None:
                                self.q_dic[q]['pos'][cur_pos] = cur_pos_nr
                                self.q_dic[q]['positions'].append(cur_pos)
                                self.q_dic['pos'][cur_pos] = q
                            # Update
                            cur_pos = col
                            cur_pos_nr = []
                        # If empty continue
                        elif len(col) == 0:
                            continue
                        else:
                            # Add numbers
                            cur_pos_nr.append(col)

                # If single
                if cur_pos not in self.q_dic[q]['positions']:
                    self.q_dic[q]['pos'][cur_pos] = cur_pos_nr
                    self.q_dic[q]['positions'].append(cur_pos)
                    self.q_dic['pos'][cur_pos] = q

            elif q_type == "NR":
                for j, line in enumerate(section):
                    if "(" in line:
                        cur_pos = "(%s)"%(line.split("(")[1].split(")")[0])
                        self.q_dic[q]['positions'].append(cur_pos)
                        self.q_dic[q]['pos']

            elif q_type == "MQ":
                pos_c = 0
                self.q_dic[q]['groups'] = pos_c
                self.q_dic[q]["pos_%i"%pos_c] = {}
                self.q_dic[q]["pos_group"] = {}
                for j, line in enumerate(section):
                    number_left_p = line.count("(")
                    if number_left_p > 1:
                        # Get next line
                        line_n = section[j+1]
                        # Get the line positions
                        line_s = line.split("|")
                        line_pos_str = line_s[1]
                        line_pos = line_pos_str.split("(")[1:]
                        line_pos_f = ["(%s)"%x.replace(" ","").replace(")","") for x in line_pos]
                        # Match
                        line_n_s = line_n.split("|")
                        line_nr_str = line_n_s[1]
                        line_nr = line_nr_str.split(" ")
                        line_nr_f = list(filter(lambda a: a != '', line_nr))
                        # Zip together
                        for cur_pos, cur_pos_nr in zip(line_pos_f, line_nr_f):
                            #self.q_dic[q]["pos_%i"%pos_c][cur_pos] = {}
                            self.q_dic[q]["pos_%i"%pos_c][cur_pos] = cur_pos_nr
                            self.q_dic[q]["pos_group"][cur_pos] = pos_c
                            self.q_dic[q]['pos'][cur_pos] = cur_pos_nr
                            self.q_dic[q]['positions'].append(cur_pos)
                            self.q_dic['pos'][cur_pos] = q
                        pos_c += 1
                        self.q_dic[q]['groups'] = pos_c
                        self.q_dic[q]["pos_%i"%pos_c] = {}

            #print(q, q_type, self.q_dic[q]['positions'])


    def find_type(self):
        # Loop over questions
        for i, q in enumerate(self.questions):
            section = self.q_dic[q]['section']
            q_type = "SQ"
            for j, line in enumerate(section):
                number_left_p = line.count("(")
                if number_left_p > 1:
                    q_type = "MQ"
                elif "|__|__|__|" in line:
                    q_type = "NR"

            # Update
            self.q_dic[q]['type'] = q_type

    def find_Q(self):
        # Make dictionary
        dic = {}
        questions = []
        # Match positions to question
        dic['pos'] = {}
        # Loop over sections
        for i, section in enumerate(self.sections):
            for j, line in enumerate(section):
                # If the Serial
                first_word = line.split(" ")[0].replace(".","")

                if "Serial number" in line:
                    pos = "(%s)"%line.split("(")[-1].split(")")[0]
                    dic['pos'][pos] = 'serial'
                    # Update
                    dic['serial'] = {} #pos
                    dic['serial']['type'] = "NR"
                    dic['serial']['pos'] = {}
                    dic['serial']['positions'] = [pos]
                    dic['serial']['col_let'] = 'A'

                    break

                # Break
                elif j == 0 and first_word not in ["ASK", "ASK:"]:
                    questions.append(first_word)
                    dic[first_word] = {}
                    dic[first_word]['section_i'] = i
                    dic[first_word]['section_j'] = j
                    dic[first_word]['section'] = section
                    break

                elif first_word in ["ASK", "ASK:"]:
                    continue

                elif "|" not in first_word:
                    questions.append(first_word)
                    dic[first_word] = {}
                    dic[first_word]['section_i'] = i
                    dic[first_word]['section_j'] = j
                    dic[first_word]['section'] = section
                    break

                elif "|" in first_word:
                    continue

                else:
                    print("Error")
                    print(line)

        #print(questions)
        self.questions = questions
        self.q_dic = dic
        # Update
        self.q_dic['questions'] = questions

    def find_sections(self):
        # Loop
        all_sections = []
        section = []
        for i, line in enumerate(self.lines):
            line = line.strip()
            # Dont look at empty lines
            if line in ['\n', '\r\n'] or len(line) == 0:
                continue

            # Start collection
            # If new section has arrived
            if "_____________________________________________" in line:
                if len(section) != 0:
                    all_sections.append(section)
                # Reset
                section = []
                continue

            # Store in section
            section.append(line)

        # Append last section
        all_sections.append(section)

        # Store sections
        self.sections = all_sections

    def read_lines(self):
        # Open file
        with open(self.filename_dst) as f:
            self.lines = f.readlines()


    def copy_file(self):
        # Define filenames
        filename_src, fileext = os.path.splitext(self.ascii_f)
        self.filename_dst = filename_src + "_" +  self.cur_time+fileext
        self.json_dst = self.filename_dst+".json"

        # New destination
        if self.ascii_f_dst != None:
            ascii_src, asciiext = os.path.splitext(self.ascii_f_dst)
            self.filename_dst = ascii_src + "_" +  self.cur_time+asciiext
            self.json_dst =self.filename_dst+".json"

        # Copy
        print(self.filename_dst)
        shutil.copy2(self.ascii_f, self.filename_dst)
