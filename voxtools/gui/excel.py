################################################################################
# Copyright (C) Voxmeter A/S - All Rights Reserved
#
# Voxmeter A/S
# Borgergade 6, 4.
# 1300 Copenhagen K
# Denmark
#
# Written by Troels Schwarz-Linnet <tsl@voxmeter.dk>, 2018
# 
# Unauthorized copying of this file, via any medium is strictly prohibited.
#
# Any use of this code is strictly unauthorized without the written consent
# by Voxmeter A/S. This code is proprietary of Voxmeter A/S.
# 
################################################################################

import sys, os, os.path

from PyQt5.QtWidgets import (QApplication,
                            QComboBox,
                            QHBoxLayout,
                            QLabel,
                            QTextEdit,
                            QListWidget, 
                            QMessageBox,
                            QVBoxLayout,
                            QWidget,
                            )
from PyQt5.QtCore import Qt, QTimer, QSize
from PyQt5.QtGui import QIcon, QPixmap

# Import openpyxl for testing
from openpyxl import load_workbook

# Import voxtools excel
from voxtools import excel, textblob_classifying, sklearn_multilabel

# Get the gui directory
gui_dir = os.path.dirname(__file__)
gui_logo = os.path.join(gui_dir, 'icons', 'gui_logo.png')
gui_logo_small = os.path.join(gui_dir, 'icons', 'gui_logo_small.png')
if os.name == 'nt': 
    WindowIcon = os.path.join(gui_dir, 'icons', 'WindowIcon_winICO.ico')
else: 
    #WindowIcon = os.path.join(gui_dir, 'icons', 'WindowIcon.svg')
    WindowIcon = os.path.join(gui_dir, 'icons', 'WindowIcon.png')
shared_data_dir = os.path.abspath(os.path.join(gui_dir, '..', 'test_suite', 'shared_data'))

def get_path(path):
    if os.name == 'nt': 
        path = path.replace('\\', '/')
    return path

# Define the texts for the combo
combo_methods = ['Prettify Tabulation report',
                        'Single label text classification',
                        'Multi label text classification']
combo_00 = \
"""Prettify Tabulation report

After a tabulation job, copy the output
from the browser into an Excel document.
Just store the whole output in 1 sheet.

Save, and then drag the Excel file here.

Example:
%s
"""%( get_path( os.path.abspath(os.path.join(shared_data_dir, 'wb05.xlsx')) ) )

combo_01 = \
"""Single label text classification

Example:
%s

Make a similar file, and then drag the Excel file here.
"""%( get_path( os.path.abspath(os.path.join(shared_data_dir, 'kodning01.xlsx')) ) )

combo_02 = \
"""Multi label text classification

Example:
%s

Make a similar file, and then drag the Excel file here.
"""%( get_path( os.path.abspath(os.path.join(shared_data_dir, 'multikodning01.xlsx')) ) )

combo_texts = [combo_00, combo_01, combo_02]

def get_icon(icon_path, resample=False):
    """Return image inside a QIcon object
    default: default image name or icon
    resample: if True, manually resample icon pixmaps for usual sizes
    (16, 24, 32, 48, 96, 128, 256). This is recommended for QMainWindow icons 
    created from SVG images on non-Windows platforms due to a Qt bug (see 
    http://code.google.com/p/spyderlib/issues/detail?id=1314)."""
    icon = QIcon( icon_path )

    if resample:
        icon0 = QIcon()
        for size in (16, 24, 32, 48, 96, 128, 256, 512):
            icon0.addPixmap(icon.pixmap(size, size))
        return icon0 
    else:
        return icon

class MainTableWidget(QWidget):

    def __init__(self, parent=None):
        super(MainTableWidget,self).__init__(parent)

        # Title
        self.setWindowTitle('Make Excel report')
        # Icon for taskbar
        if os.name == 'nt':
            self.setWindowIcon(QIcon( get_path(WindowIcon) ))
        else:
            resample = os.name != 'nt'
            icon = get_icon( get_path(WindowIcon), resample=resample)
            #print(dir(icon))
            #print(icon.availableSizes())
            self.setWindowIcon( icon )

        # Size and position
        w = 350; h=800
        self.resize(w, h)
        self.move(0, 0)

        # Make overall layout
        widgetLayout = QVBoxLayout(self)

        # Create layout and widgets for methods
        widgetLayout_methods = QHBoxLayout()
        # Make label and combo widget
        methods_lbl = QLabel('Chose method:')

        # Create and fill the combo box to choose the method
        self.method_combo = QComboBox()
        self.method_combo.addItems(combo_methods)
        # Add connect
        self.method_index = 0
        self.method_combo.currentIndexChanged.connect(self.method_change)

        # Add widgets to methods
        widgetLayout_methods.addWidget(methods_lbl)
        widgetLayout_methods.addWidget(self.method_combo)
        #widgetLayout_methods.addStretch()
        # Add to vertical
        widgetLayout.addLayout(widgetLayout_methods)

        # Create layout and widgets for info
        widgetLayout_info_lbl = QHBoxLayout()
        # The label for the info widget
        info_lbl = QLabel('Info:')
        # Add widgets to info
        widgetLayout_info_lbl.addWidget(info_lbl)
        # Add to vertical
        widgetLayout.addLayout(widgetLayout_info_lbl)
        
        # Create layout and widget for info
        widgetLayout_info = QVBoxLayout()
        # The textbox with info
        self.info_textbox = QTextEdit()
        self.info_textbox.setFixedSize(QSize(w-20, 150))
        # Set Read only
        self.info_textbox.setReadOnly(True)
        #self.info_textbox.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        # Set text
        self.set_info_in_textbox(combo_texts[self.method_index])
        # Add widgets to info
        widgetLayout_info.addWidget(self.info_textbox)
        #widgetLayout_info.addStretch()
        # Add to vertical
        widgetLayout.addLayout(widgetLayout_info)

        # Make drop_box
        widgetLayout_drop_box = QVBoxLayout()
        # Create drop_box
        drop_box_lbl = QLabel('Drag & Drop files to the box below:')
        self.drop_box = TestListBox()
        # Add widget
        widgetLayout_drop_box.addWidget(drop_box_lbl)
        widgetLayout_drop_box.addWidget(self.drop_box)
        # Add to vertical
        widgetLayout.addLayout(widgetLayout_drop_box)
        #widgetLayout_drop_box.addStretch()        
        
        # Make layout for whole gui
        self.setLayout(widgetLayout)  

    def method_change(self, i):
        # Store the index
        self.method_index = i
        # Set text
        self.set_info_in_textbox(combo_texts[self.method_index])

        # Set method to drop_box
        self.drop_box.method_index = i
        
        #print("Items in the list are :")
        #for j in range(self.method_combo.count()):
        #    print(j, self.method_combo.itemText(j))
        #print("Current index '", i, "' selection changed '", self.method_combo.currentText(), "'")

    def set_info_in_textbox(self, text):
        self.info_textbox.setText(text)


class TestListBox(QListWidget):
    def __init__(self, parent=None):
        super(TestListBox, self).__init__(parent)
        self.setAcceptDrops(True)

        # Set background
        self.setStyleSheet("""QListWidget {
                    background-color: white;
                    background-image: url(%s);
                    background-position: center;
                    background-repeat: no-repeat;
                    }"""%get_path(gui_logo_small))

        # The initial method index
        self.method_index = 0

    # http://pyqt.sourceforge.net/Docs/PyQt5/api/QtGui/qdragenterevent.html
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls:
              event.accept()
        else:
              event.ignore()

    # http://pyqt.sourceforge.net/Docs/PyQt5/api/QtGui/qdragmoveevent.html
    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls:
              event.setDropAction(Qt.CopyAction)
              event.accept()
        else:
              event.ignore()

    # http://pyqt.sourceforge.net/Docs/PyQt5/api/QtGui/qdropevent.html
    def dropEvent(self, event):
        if event.mimeData().hasUrls:
            event.setDropAction(Qt.CopyAction)
            event.accept()
            # Get the urls
            all_urls = event.mimeData().urls()

            # Prettify Tabulation report
            success = False
            if self.method_index == 0:
                success = self.execute_excel_urls(all_urls=all_urls, method_index=self.method_index)
            # Single label text classification
            elif self.method_index == 1:
                success = self.execute_excel_urls(all_urls=all_urls, method_index=self.method_index)
            # Multi label text classification
            elif self.method_index == 2:
                success = self.execute_excel_urls(all_urls=all_urls, method_index=self.method_index)

            # Show dialog
            if success:
                self.ok_dialog(timeout=3, text="All files has been converted!")

        else:
            event.ignore()

    def execute_excel_urls(self, all_urls=[], method_index=0):
        # Loop over urls passed
        success = False
        for url in all_urls:
            # Possible convert
            if str(type(url)) == "<class 'PyQt5.QtCore.QUrl'>":
                link_str = str(url.toLocalFile())
            else:
                link_str = url
            # Get filename_src
            filename_src, fileext = os.path.splitext(link_str)
            # If not Excel
            if fileext != '.xlsx':
                print("Not Excel file! Skipping file: %s" % link_str)
                self.showdialog(Text="Not an Excel file", 
                                InformativeText="Please provide an Excel file for method: '%s'"%combo_methods[method_index],
                                DetailedText="File: %s"%link_str)
                continue

            # If Excel
            # Try to autodetect the mode
            wb = load_workbook(link_str)
            ws = wb.active
            ws_A1 = ws['A1']

            # Detect if 'Multi label text classification'
            if 'target_categories' in wb.sheetnames:
                method_index_detect = 2
            # Detect if 'Single label text classification'
            elif ws_A1.value == 'sentences': 
                method_index_detect = 1
            # Else expect 'Prettify Tabulation report'
            else:
                method_index_detect = 0

            # Now test
            if method_index != method_index_detect:
                self.showdialog(Text="Mismatch in method chosen and detected file", 
                                InformativeText="Selected method: '%s' \n\nDetected method: '%s' " % (combo_methods[method_index], combo_methods[method_index_detect]) ,
                                DetailedText="File: %s"%link_str)
                continue

            # 'Prettify Tabulation report'
            if method_index == 0:
                # Instantiate the Excel class and run it
                exl = excel.excel(excel_src=link_str)
                exl.run_all()

            # 'Single label text classification'
            elif method_index == 1:
                # Instantiate the Textblob class and run it
                tcl = textblob_classifying.text(excel_src=link_str)
                tcl.run_all()

            # 'Multi label text classification'
            elif method_index == 2:
                # Instantiate the Textblob class and run it
                tcl =  sklearn_multilabel.text(excel_src=link_str)
                tcl.run_all()

            # Update success
            success = True

        return success
    
    def ok_dialog(self, timeout=3, text=None):
        messagebox = TimerMessageBox(timeout=timeout, text=text, parent=self)
        messagebox.exec_()

    def showdialog(self, Text="This is a message box", InformativeText="This is additional information", DetailedText=None):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)

        # Set Title
        msg.setWindowTitle("Message")

        # Set text
        msg.setText(Text)
        msg.setInformativeText(InformativeText)

        # Set additional info
        if DetailedText != None:
            msg.setDetailedText(DetailedText)

        # Set button and action
        #msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.buttonClicked.connect(self.msgbtn)

        retval = msg.exec_()
        #print("value of pressed message box button:", retval)

    def msgbtn(self,i):
        pass
        #print("Button pressed is:",i.text())

class TimerMessageBox(QMessageBox):
    def __init__(self, timeout=3, text=None, parent=None):
        super(TimerMessageBox, self).__init__(parent)
        self.setWindowTitle("Ok status")
        self.time_to_wait = timeout
        self.text = text
        if self.text == None:
            self.setText("wait (closing automatically in {0} secondes.)".format(timeout))
        else:
            self.setText("{0}\n\n      Window closes in {1}s".format(self.text, self.time_to_wait))
        self.setStandardButtons(QMessageBox.NoButton)
        self.timer = QTimer(self)
        self.timer.setInterval(1000)
        self.timer.timeout.connect(self.changeContent)
        self.timer.start()

    def changeContent(self):
        if self.text == None:
            self.setText("wait (closing automatically in {0} secondes.)".format(self.time_to_wait))
        else:
            self.setText("{0}\n\n      Window closes in {1}s".format(self.text, self.time_to_wait))
        self.time_to_wait -= 1
        if self.time_to_wait <= 0:
            self.close()

    def closeEvent(self, event):
        self.timer.stop()
        event.accept()

if __name__ == '__main__':

    app = QApplication(sys.argv)
    #ex = TestListBox()
    ex = MainTableWidget()
    ex.show()
    app.exec_()
