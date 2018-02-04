################################################################################
# Copyright (C) Voxmeter A/S - All Rights Reserved
#
# Voxmeter A/S
# Borgergade 6, 4.
# 1300 Copenhagen K
# Denmark
#
# Written by Troels Schwarz-Linnet <tsl@voxmeter.dk>, 2018
# 
# Unauthorized copying of this file, via any medium is strictly prohibited.
#
# Any use of this code is strictly unauthorized without the written consent
# by Voxmeter A/S. This code is proprietary of Voxmeter A/S.
# 
################################################################################
# Inspiration: 
# http://scikit-learn.org/dev/modules/multiclass.html
# https://stackoverflow.com/questions/10526579/use-scikit-learn-to-classify-into-multiple-categories
# https://stackoverflow.com/questions/10526579/use-scikit-learn-to-classify-into-multiple-categories/19172087#19172087

import shutil, datetime, os, os.path, sys
from distutils.version import StrictVersion
from operator import itemgetter 

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Font, PatternFill, Border

# Check version
from openpyxl import __version__
test_version = StrictVersion(__version__) < StrictVersion("2.5.0")
if test_version:
    print("You need to have openpyxl version 2.5.0. You have %s"%__version__)
    sys.exit()
    
import numpy as np
from sklearn.pipeline import Pipeline
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.svm import LinearSVC
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.multiclass import OneVsRestClassifier
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.metrics import precision_score

class text:
    def __init__(self, excel_src=None, excel_dst=None):
        # Store 
        self.excel_src = excel_src
        self.excel_dst = excel_dst

        # Make current time
        #self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W-H%H-M%M-S%S")
        self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W")

    def run_all(self):
        # Copy Excel workbook
        self.copy_excel()
        # Print all objects of workbook
        #[print(i) for i in dir(self.wb)]
        #print(self.wb.active, self.wb.properties)

        # Open workbook
        self.load_wb()

        # Delete empty sheets
        self.delete_empty_sheets()

        # Read data
        self.read_data()

        # Read data
        self.read_target_categories()

        # Now train
        self.do_train()

        # Now classify all sentences
        self.do_classify_sentences()

        # Now write to worksheet
        self.write_results()

        # Save Workbook
        self.wb.save(self.excel_dst)


    def write_results(self):
        # Loop over sentences
        sent_irows, sentences = self.sent_dat
        for i, sentence in enumerate(sentences):
            # Extract and unpack
            sent_irow = sent_irows[i]
            iRow, iCol = sent_irow
            # Extract and unpack
            sent_classification = self.sent_classifications[iRow-1]
            classi_sent, classi_class = sent_classification
            classi_class_str = ' '.join(map(str, classi_class))
            # Sanity check
            if sentence != classi_sent:
                print("Oh no!!! Error! %s != %s" %(sentence, classi_sent) )
                continue

            # Get the cell
            cCell = self.ws['%s%i'%(self.index_class_L, iRow+1)]
            # Store the classification
            cCell.value = classi_class_str

        # Loop over train
        train_irows, train_sentences = self.train_dat
        correct_train = []
        for i, sentence_pair in enumerate(train_sentences):
            sentence, classi_train = sentence_pair
            classi_train_set = self.create_set_from_str(classi_train)
            classi_train_cat = self.create_cat_from_str(classi_train)
            # Extract and unpack
            train_irow = train_irows[i]
            iRow, iCol_1, iCol_2 = train_irow
            # Extract and unpack
            sent_classification = self.sent_classifications[iRow-1]
            classi_sent, classi_class = sent_classification
            classi_class_str = ' '.join(map(str, classi_class))
            classi_class_str_set = self.create_set_from_str(classi_class_str)
            classi_class_cat = self.create_cat_from_str(classi_class_str)
            # Sanity check
            if sentence != classi_sent:
                print("Oh no!!! Error! %s != %s" %(sentence, classi_sent) )
                continue
            # Check
            if classi_train_set == classi_class_str_set:
                equal = 1
            else:
                equal = 0
            # Add
            correct_train.append(equal)
            # Get the cell
            cCell = self.ws['%s%i'%(self.index_correct_train_L, iRow+1)]
            # Store the classification
            cCell.value = equal
            # Store the strings
            cCell = self.ws['G%i'%(iRow+1)]
            cCell.value = classi_train_cat
            cCell = self.ws['H%i'%(iRow+1)]
            cCell.value = classi_class_cat


        # Write results
        self.ws['J1'].value = "correct_train"
        self.ws['J2'].value = "sum"
        self.ws['K2'].value = sum(correct_train)
        self.ws['J3'].value = "total"
        self.ws['K3'].value = len(correct_train)
        self.ws['J4'].value = "pct"
        self.ws['K4'].value = float("%0.2f"%( sum(correct_train)/len(correct_train) ))

        # Loop over test
        correct_test = []
        if self.has_test:
            test_irows, test_sentences = self.test_dat
            for i, sentence_pair in enumerate(test_sentences):
                sentence, classi_test = sentence_pair
                classi_test_set = self.create_set_from_str(classi_test)
                classi_test_cat = self.create_cat_from_str(classi_test)
                # Extract and unpack
                test_irow = test_irows[i]
                iRow, iCol_1, iCol_2 = test_irow
                # Extract and unpack
                sent_classification = self.sent_classifications[iRow-1]
                classi_sent, classi_class = sent_classification
                classi_class_str = ' '.join(map(str, classi_class))
                classi_class_str_set = self.create_set_from_str(classi_class_str)
                classi_class_cat = self.create_cat_from_str(classi_class_str)

                # Sanity check
                if sentence != classi_sent:
                    print("Oh no!!! Error! %s != %s" %(sentence, classi_sent) )
                    continue
                # Check
                if classi_test_set == classi_class_str_set:
                    equal = 1
                else:
                    equal = 0
                # Add
                correct_test.append(equal)
                # Get the cell
                cCell = self.ws['%s%i'%(self.index_correct_test_L, iRow+1)]
                # Store the classification
                cCell.value = equal        
                # Store the strings
                cCell = self.ws['I%i'%(iRow+1)]
                cCell.value = classi_test_cat
                cCell = self.ws['H%i'%(iRow+1)]
                cCell.value = classi_class_cat


            # Write results
            self.ws['J7'].value = "correct_test"
            self.ws['J8'].value = "sum"
            self.ws['K8'].value = sum(correct_test)
            self.ws['J9'].value = "total"
            self.ws['K9'].value = len(correct_test)
            self.ws['J10'].value = "pct"
            self.ws['K10'].value = float("%0.2f"%( sum(correct_test)/len(correct_test) ))
                    
    def create_cat_from_str(self, in_str):
        if isinstance(in_str, int):
            cur_index = [in_str]
        else:
            cur_index = [*map(int, in_str.split(" "))]

        cur_categories = itemgetter(*cur_index)(self.categories[1])
        if isinstance(cur_categories, tuple):
            cur_categories_str = ','.join(map(str, cur_categories))
        else:
            cur_categories_str = cur_categories
        return cur_categories_str

    def create_set_from_str(self, in_str):
        if isinstance(in_str, int):
            cur_set = set([in_str])
        else:
            list_class = [*map(int, in_str.split(" "))]
            cur_set = set(list_class)
        return cur_set

    def do_classify_sentences(self):
        # Get the X, the sen
        X_sentences = self.sent_dat[1]

        # Predict
        predict = self.classifier.predict(X_sentences)
        predict_inv = self.mlb.inverse_transform(predict)

        # Loop over all sentences
        self.sent_classifications = []
        for sent, classification in zip(X_sentences, predict_inv):
            # Store
            self.sent_classifications.append((sent, classification))

    def do_train(self):
        # Get the X and y of problem
        train = self.train_dat[1]
        X_train, y_train = self.get_X_y(train)

        # Get the X and y of test
        if self.has_test:
            test = self.test_dat[1]
            X_test, y_test = self.get_X_y(test)

        # Create classifier
        self.mlb = MultiLabelBinarizer()
        Y = self.mlb.fit_transform(y_train)

        # Define methods to classifier
        self.classifier = Pipeline([
            ('vectorizer', CountVectorizer()),
            #('vectorizer', CountVectorizer(ngram_range=(self.vectorizer_min , self.vectorizer_max))),
            ('tfidf', TfidfTransformer()),
            ('clf', OneVsRestClassifier(LinearSVC()))])

        # Fit to classifier
        self.classifier.fit(X_train, Y)

        # Predict on test
        if self.has_test:
            # Get the right format
            y_true = self.mlb.fit_transform(y_test)
            y_pred = self.classifier.predict(X_test)
            # Get labels
            y_pred_labels = self.mlb.inverse_transform(y_pred)
            # 'macro': Calculate metrics for each label, and find their unweighted mean. This does not take label imbalance into account.
            # 'micro': Calculate metrics globally by counting the total true positives, false negatives and false positives.
            self.accuracy_test = precision_score(y_true, y_pred, average='micro')
            
            print("After training %i sentences, the accuracy on %i test sentences is %0.2f"% (len(train), len(test), self.accuracy_test) )

    def get_X_y(self, categories):
        # Convert
        X_arr = []
        y_arr = []

        for sent, cat in categories:
            X_arr.append(sent)
            # Split on space
            if isinstance(cat, int):
                y = [cat]
            elif " " in cat:
                cat_split = cat.split(" ")
                y = []
                for cat_i in cat_split:
                    y.append(int(cat_i))
            else:
                y = [int(cat)]
            # Add
            y_arr.append(y)

        return(X_arr, y_arr)

    def read_target_categories(self):
        # Use the first sheet
        current_ws = self.ws
        ws_name = "target_categories"
        self.ws = self.wb[ws_name]            

        # Get row 1
        self.header_target_vals = []
        for iCol in range(1, self.ws.max_column+1):
            iCol_L = get_column_letter(iCol)
            cCell = self.ws['%s1'%(iCol_L)]
            self.header_target_vals.append(cCell.value)
        # Check for header
        if 'categories' in self.header_target_vals:
            self.has_header_target = True
            self.iRow_target_skip = 1
        else:
            self.has_header_target = False
            self.header_target_vals = ['index', 'categories', 'vectorizer_min', 'vectorizer_max']
            self.iRow_target_skip = 0

        # Get the index of columns
        if 'index' in self.header_target_vals:
            self.index_target_index = self.header_target_vals.index('index')
        else:
            self.index_target_index = 0

        if 'categories' in self.header_target_vals:
            self.index_categories = self.header_target_vals.index('categories')
        else:
            self.index_categories = 1

        if 'vectorizer_min' in self.header_target_vals:
            self.index_vectorizer_min = self.header_target_vals.index('vectorizer_min')
        else:
            self.index_vectorizer_min = 2
        self.index_vectorizer_min_L = get_column_letter(self.index_vectorizer_min+1)

        if 'vectorizer_max' in self.header_target_vals:
            self.index_vectorizer_max = self.header_target_vals.index('vectorizer_max')
        else:
            self.index_vectorizer_max = 3
        self.index_vectorizer_max_L = get_column_letter(self.index_vectorizer_max+1)

        # Collect all categories
        self.categories = self.collect_column_data(self.iRow_target_skip, self.index_categories)

        # Collect 
        self.vectorizer_min = int(self.ws['%s%i'%(self.index_vectorizer_min_L, 2)].value)
        self.vectorizer_max = int(self.ws['%s%i'%(self.index_vectorizer_max_L, 2)].value)

        # Go back
        self.ws = current_ws

    def read_data(self):
        # Use the first sheet
        ws_name = self.wb.sheetnames[0]
        self.ws = self.wb[ws_name]

        # Get row 1
        self.header_vals = []
        for iCol in range(1, self.ws.max_column+1):
            iCol_L = get_column_letter(iCol)
            cCell = self.ws['%s1'%(iCol_L)]
            self.header_vals.append(cCell.value)
        # Check for header
        if 'train' in self.header_vals:
            self.has_header = True
            self.iRow_skip = 1
        else:
            self.has_header = False
            self.header_vals = ['sentences', 'train', 'classify', 'test', 'correct_train', 'correct_test']
            self.iRow_skip = 0

        # Get the index of columns
        if 'sentences' in self.header_vals:
            self.index_sent = self.header_vals.index('sentences')
        else:
            self.index_sent = 0

        if 'train' in self.header_vals:
            self.index_train = self.header_vals.index('train')
        else:
            self.index_train = 1

        if 'classify' in self.header_vals:
            self.index_class = self.header_vals.index('classify')
        else:
            self.index_class = 2
        self.index_class_L = get_column_letter(self.index_class+1)

        if 'test' in self.header_vals:
            self.index_test = self.header_vals.index('test')
            self.has_test = True
        else:
            self.index_test = 3
            self.has_test = False

        if 'correct_train' in self.header_vals:
            self.index_correct_train = self.header_vals.index('correct_train')
        else:
            self.index_correct_train = 4
        self.index_correct_train_L = get_column_letter(self.index_correct_train+1)

        if 'correct_test' in self.header_vals:
            self.index_correct_test = self.header_vals.index('correct_test')
        else:
            self.index_correct_test = 5
        self.index_correct_test_L = get_column_letter(self.index_correct_test+1)

        # Collect all sentences
        self.sent_dat = self.collect_column_data(self.iRow_skip, self.index_sent)

        # Collect paired values of sentences and train
        self.train_dat = self.collect_column_data(self.iRow_skip, self.index_sent, self.index_train)

        # If test is available
        if self.has_test:
            self.test_dat = self.collect_column_data(self.iRow_skip, self.index_sent, self.index_test)

    def collect_column_data(self, iRow_skip=0, col_i_1=1, col_i_2=None):
        # Loop over rows
        iRow_nrs = []
        iRow_pair_vals = []
        
        for iRow, cCol in enumerate(self.ws.rows):
            # Possible skip
            if iRow < iRow_skip:
                continue
            if col_i_2 != None:
                # Get the cells
                cCell_1 = cCol[col_i_1]
                cCell_2 = cCol[col_i_2]
                # Get the value
                val_1 = cCell_1.value
                val_2 = cCell_2.value
                if val_1 != None and val_2 != None:
                    iRow_nrs.append((iRow,col_i_1,col_i_2))
                    iRow_pair_vals.append((val_1, val_2))

            elif col_i_2 == None:
                # Get the cells
                cCell_1 = cCol[col_i_1]
                # Get the value
                val_1 = cCell_1.value
                if val_1 != None:
                    iRow_nrs.append((iRow,col_i_1))
                    iRow_pair_vals.append(val_1)

        if len(iRow_nrs) > 0:
            return [iRow_nrs, iRow_pair_vals]
        else:
            return None

    def delete_empty_sheets(self):
        # Loop through worksheets
        for ws in self.wb:
            #print(ws.title, ws.max_row, ws.max_column)
            # Delete sheet if empty
            if ws.max_row == 1 and ws.max_column == 1:
                #print(ws.title)
                std=self.wb[ws.title]
                self.wb.remove(std)


    def load_wb(self):
        # Open workbook
        self.wb = load_workbook(self.excel_dst, data_only=True)


    def copy_excel(self):
        filename_src, fileext = os.path.splitext(self.excel_src)
        filename_dst = filename_src + "_" +  self.cur_time
        # New destination
        if self.excel_dst == None:
            # If this is an update on file with date in it
            if self.cur_time in filename_src:
                filename_split = filename_src.split(self.cur_time)
                if filename_split[-1] == "":
                    filename_dst = filename_split[0] + self.cur_time + "_002"
                else:
                    version = int(filename_split[-1].split("_")[-1])
                    version += 1
                    filename_dst = filename_split[0] + self.cur_time + "_%03d"%(version)
            
            # Create the file
            self.excel_dst = filename_dst+fileext

        # Copy, but do not overwrite !
        if os.path.isfile(self.excel_dst):
            print("File %s does already exists! Exiting!"%(self.excel_dst))
            sys.exit()
        else:
            shutil.copy2(self.excel_src, self.excel_dst)