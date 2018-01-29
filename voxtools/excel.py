################################################################################
# Copyright (C) Troels Schwarz-Linnet - All Rights Reserved
# Written by Troels Schwarz-Linnet <tlinnet@gmail.com>, January 2018
# 
# Unauthorized copying of this file, via any medium is strictly prohibited.
#
# Any use of this code is strictly unauthorized without the written consent
# by the the author. This code is proprietary of the author.
# 
################################################################################
import shutil, datetime, os.path, copy, operator, math, sys
from distutils.version import StrictVersion

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Font, PatternFill, Border

# Print version
from openpyxl import __version__
test_version = StrictVersion(__version__) < StrictVersion("2.5.0")
if test_version:
    print("You need to have openpyxl version 2.5.0. You have %s"%__version__)
    sys.exit()

# Set standards
SKIP_THESE = [
    '[+] og [-] = signifikant forskel på 95% eller mere',
]
Col_width_max = 177 # 100+50+20+5+2

def calc_text_width(cCell):
    value = cCell.value
    if value is None:
        length = 0
        return length
    # Get string length
    str_length = len(str(value))
    # Get number format
    number_format = cCell.number_format
    # Get font names
    font_name = cCell.font.name
    # Get font size
    font_sz = cCell.font.sz
    # Get if bold
    font_b = cCell.font.b

    # In font_name is Arial
    if font_name == 'Arial':
        # This has been manually found in Excel
        if font_sz == 5.0:
            length = str_length * 0.35375 + 1.64625
        # This is the summary box
        elif font_sz == 11.0 and font_b:
            # First convert to general
            if number_format == '0.0%':
                val_str = "{0:.1f}%".format(value * 100)
                str_length = str_length = len(val_str)
                #length = str_length * 1.00 + 2.
                length = str_length * 1.00 + 2.2
            else:
                # length = str_length * 1.15 + 1.58
                length = str_length * 1.15 + 1.8
        # This is the absolut and percent area
        elif font_sz == 9.0 and not font_b:
            if number_format == '0.0%':
                val_str = "{0:.1f}%".format(value * 100)
                str_length = str_length = len(val_str)
                #length = str_length * 0.835 + 1.989
                length = str_length * 0.84 + 2.
            else:
                # length = str_length * 0.6845 + 1.856
                length = str_length * 0.7 + 2.1
        # This is everything else
        else:
            print("Arial, but unknown fontsize: %i  %s"%(font_sz, value))
            length = str_length

        # Finally return
        length_round = float("{0:.2f}".format(length))
        return length_round
    # If Verdana
    elif font_name == 'Verdana':
        if value != ' Luk':
            print("Verdana: '%s'"%value)
        return str_length
    # Else
    else:
        print("Font type is: %s . I do not know length."%font_name)
        return str_length

class excel:
    def __init__(self, excel_src=None, excel_dst=None, debug_sheet=False):
        # Store 
        self.excel_src = excel_src
        self.excel_dst = excel_dst
        self.debug_sheet = debug_sheet

        # Make current time
        #self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W-H%H-M%M-S%S")
        self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W")


    def run_all(self):
        # Copy Excel workbook
        self.copy_excel()
        # Print all objects of workbook
        #[print(i) for i in dir(self.wb)]
        #print(self.wb.active, self.wb.properties)

        # Open workbook
        self.load_wb()

        # Delete empty sheets
        self.delete_empty_sheets()

        # Format number percent format per cell
        self.Format_Cells_number_format()

        # Find the Sheet groups
        self.Find_sheet_groups()

        # Copy groups to sheets
        ###self.Copy_groups_to_sheets_as_new()
        self.Copy_groups_to_sheets_as_copy()

        # Set width of cells
        self.Format_Column_Width()

        # Set Page Layout per sheet
        self.Page_Layout_setup()

        # Set height of cells
        self.Format_Column_Height()

        # Delete rows
        #self.Worksheet_delete_rows()

        # Save Workbook
        self.wb.save(self.excel_dst)

        print("Done with %s"%self.excel_dst)

    def Worksheet_delete_rows(self):
        for ws in self.wb:
            # Loop over rows
            freq_groups = []
            cont_while = True
            while cont_while:
                iRow_prev = None
                delete_start = None
                delete_end = None
                for iRow, cCol in enumerate(ws.rows):
                    if ws.row_dimensions[iRow].height == 0.1:
                        # Update initial
                        #print(ws.title, iRow, iRow_prev, ws.max_row)
                        if iRow_prev == None:
                            iRow_prev = iRow
                            delete_start = iRow
                        # If they are equal in the start
                        elif iRow == iRow_prev:
                            continue
                        # If reached the end
                        elif iRow == (ws.max_row - 1):
                            delete_end = iRow
                        # If the difference is expected 1, then update
                        elif iRow - iRow_prev == 1:
                            iRow_prev = iRow
                            delete_end = iRow
                        elif iRow - iRow_prev > 1:
                            delete_end = iRow_prev
                        #print(t)
                if delete_start != None and delete_end != None:
                    cont_while = True
                    amount = delete_end-delete_start
                    #print(ws.title, delete_start, amount)
                    #ws.delete_rows(delete_start, amount)
                    ws.delete_rows(delete_start)
                else:
                    cont_while = False
                #print()
                #print(ws.title, delete_start, delete_end, cont_while)

    def Format_Column_Height(self):
        # Loop through
        for sheet_group in self.sheet_groups_dict['sheetname_new']:
            groupname, cell_from, cell_to, sheetname_new, group_questions, group_questions_coord, col_titles, col_widths, col_widths_info = sheet_group
            # Get ws
            ws = self.wb[sheetname_new]

            # Group row lists
            group_row_range = range(cell_from[1], cell_to[1]+1)
            #print(group_row_range)

            # Loop over rows
            for iRow, cCol in enumerate(ws.rows):
                iRow += 1
                # If row in group range
                if iRow in group_row_range:
                    # Loop over columns
                    col_title_heights = []
                    col_title_heights_iRow = None
                    for iCol, cCell in enumerate(cCol):
                        iCol += 1
                        iCol_L = get_column_letter(iCol)
                        # Get the width of column
                        iCol_w = ws.column_dimensions[iCol_L].width
                        iCol_w_N_Sum = iCol_w

                        # Run over neighbour columns, and collect their widths
                        # This is to check if cells are concatinated
                        for iCol_N in range(iCol+1, iCol+3):
                            # Get the letter
                            iCol_N_L = get_column_letter(iCol_N)
                            # Get the cell
                            cCell_N = ws['%s%i'%(iCol_N_L, iRow)]
                            if cCell_N.value == None:
                                iCol_w_N = ws.column_dimensions[iCol_N_L].width
                                if iCol_w_N != None:
                                    iCol_w_N_Sum += iCol_w_N
                            else:
                                break

                        # If the Groupname, set fixed
                        #if cCell.value == groupname or (cCell.value == "(abs)" or cCell.value == "(procent)"):
                        if cCell.value == groupname:
                            ws.row_dimensions[iRow].height = 15

                        # Set fix for the small abs / percent boxes
                        if cCell.font.sz == 5.0:
                            ws.row_dimensions[iRow].height = 16

                        # First set fixed height on group questions
                        if cCell.value in group_questions:
                            #print(iRow, iCol, cCell.value)
                            ws.row_dimensions[iRow].height = 30
                            #print(iRow, cCell.value)

                        # If the col_titles, collect the
                        if cCell.value in col_titles:
                            # Get the length
                            length = calc_text_width(cCell)
                            # Divide with with, try with including neighbour widt
                            #w_l_div = length / iCol_w
                            w_l_div = length / iCol_w_N_Sum
                            # Fraction of division
                            w_l_div_c = math.ceil(w_l_div)
                            # Each line should get 15 pt
                            col_title_h = 15.0 * w_l_div_c
                            #print(length, iCol_w, iCol_w_N_Sum, col_title_h, cCell.value)
                            # Store the width
                            col_title_heights.append(col_title_h)
                            # Store the line number
                            col_title_heights_iRow = iRow

                    # After running through the collumns
                    # Set to the maximum of the collected col_title heights
                    if len(col_title_heights) != 0:
                        #print(col_title_heights, cCell.value, iCol, iCol_L, iRow, col_title_heights_iRow, max(col_title_heights))
                        #print(max(col_title_heights), iRow, iCol, sheetname_new, col_title_heights)
                        ws.row_dimensions[iRow].height = max(col_title_heights)

                    # Now check for the rest
                    cCell = ws['A%i'%(iRow)]
                    if cCell.font.sz == 9.0:
                        # Get the width
                        iCol_w = ws.column_dimensions['A'].width
                        # Get the length
                        length = calc_text_width(cCell)
                        # Divide with with
                        w_l_div = length / iCol_w + 0.45
                        # Fraction of division
                        w_l_div_c = math.ceil(w_l_div)
                        # Each line should get 15 pt
                        row_h = 15.0 * w_l_div_c
                        # Set height
                        #print(w_l_div_c, w_l_div, cCell.value)
                        #print(iCol_w, length, row_h, w_l_div, cCell.value)
                        ws.row_dimensions[iRow].height = row_h

                # First and second row should not be changed.
                # This is so we can easily delete the empty rows
                elif iRow == 1 or iRow == group_row_range[0] - 1:
                    continue
                # Since we cannot delete rows, then at least try to hide them.
                # elif iRow < group_row_range[0]:
                else:
                    ws.row_dimensions[cCol[0].row].height = 0.1

    def Format_Column_Width(self):

        # Loop through
        for i, sheet_group in enumerate(self.sheet_groups_dict['sheetname_new']):
            groupname, cell_from, cell_to, sheetname_new = sheet_group
            # Get ws
            ws = self.wb[sheetname_new]

            col_titles = []
            col_widths = []
            col_widths_sum = 0.0
            group_questions = []
            group_questions_coord = []

            # Loop over columns
            for iCol, cCol in enumerate(ws.columns):
                iCol += 1

                cCol_lengts = []
                cCol_lengts_info = []
                for iRow, cCell in enumerate(cCol):
                    iRow += 1

                    # If first column, continue to next column
                    if iCol == 1:
                        break

                    # Collect questions for groups
                    elif cCell.value != None and (cCell.value == "(abs)" or cCell.value == "(procent)"):
                        # Get the question from column A
                        question = ws['A%i'%iRow].value
                        if question != None:
                            group_questions.append(question)
                            group_questions_coord.append(cCell.coordinate)

                    # Collect questions for groups
                    elif cCell.value != None and cCell.value == groupname:
                        # Get the question from column A
                        question = ws['A%i'%iRow].value
                        if question != None:
                            group_questions.append(question)
                            group_questions_coord.append(cCell.coordinate)

                    # Do not determine width, if None, and not the group name
                    elif cCell.value != None and cCell.value != groupname and cCell.value not in SKIP_THESE:
                        # Collect headers to skip
                        if iRow == cell_from[1]+1:
                            # And not already collected
                            if cCell.value not in col_titles:
                                col_titles.append(cCell.value)

                        # Skip titles, but else calculate
                        if cCell.value not in col_titles:
                            length = calc_text_width(cCell)
                            cCol_lengts.append(length)
                            cCol_lengts_info.append((cCell.coordinate, cCell.value, length))

                # Get max length of column
                if len(cCol_lengts) != 0:
                    cCol_lengt_max_index, cCol_lengt_max = max(enumerate(cCol_lengts), key=operator.itemgetter(1))
                    # For weird reasons. The width should be added 0.83 everytime
                    cCol_lengt_max_use = cCol_lengt_max + 0.83
                    ws.column_dimensions[cCell.column].width = cCol_lengt_max_use

                    # Store column col_widths
                    col_widths.append((iCol, get_column_letter(iCol), cCol_lengt_max, cCol_lengt_max_use))
                    # Collect sum
                    #col_widths_sum += cCol_lengt_max_use
                    col_widths_sum += cCol_lengt_max

                    #print(sheetname_new, cCol_lengts_info[cCol_lengt_max_index])
                    #print(cCol_lengt_max)

            # Add widths per new_sheet_name
            self.sheet_groups_dict['sheetname_new'][i].append(group_questions)
            self.sheet_groups_dict['sheetname_new'][i].append(group_questions_coord)
            self.sheet_groups_dict['sheetname_new'][i].append(col_titles)
            self.sheet_groups_dict['sheetname_new'][i].append(col_widths)

            # Set width of column A
            col_widths_sum = float("{0:.2f}".format(col_widths_sum))
            col_width_A_calc = Col_width_max - col_widths_sum
            col_width_A_calc = float("{0:.2f}".format(col_width_A_calc))
            # If negative, set standard
            if col_width_A_calc < 0.0:
                col_width_A = 15
            elif col_width_A_calc > 40.0 :
                col_width_A = 40.
            else:
                col_width_A = col_width_A_calc
            # For weird reasons. The width should be added 0.83 everytime
            col_width_A = float("{0:.2f}".format(col_width_A+0.83))
            # Set width of column A
            ws.column_dimensions["A"].width = col_width_A
            # Collect
            self.sheet_groups_dict['sheetname_new'][i].append((Col_width_max, col_widths_sum, col_width_A))
            if self.debug_sheet:
                print("Col width A: ", "'%s'"%sheetname_new, (Col_width_max, col_widths_sum, col_width_A_calc))

        # Print last info
        #print(self.sheet_groups_dict['sheetname_new'])
        #print(col_widths_sum)

    def Copy_groups_to_sheets_as_copy(self):
        # Get the list of current sheet names
        wb_cur_sheets = self.wb.sheetnames
        # Collect new target sheetnames
        self.sheet_groups_dict['sheetname_new'] = []
        sheetname_new_list = []

        # Loop through worksheets
        for ws in self.wb:
            keys = self.sheet_groups_dict[ws.title]['keys']
            if self.debug_sheet:
                #keys = keys[:1]
                keys = [keys[0]]+[keys[4]]
            for key in keys:
                # Extract
                groupname, cell_from, cell_to = self.sheet_groups_dict[ws.title][key]
                # Define new sheetname
                if groupname in wb_cur_sheets:
                    sheetname_new = ws.title + "_" + groupname
                else:
                    sheetname_new = groupname
                # Replace bad characters. Max 31 characters in Excel.
                sheetname_new = sheetname_new.replace("/","")[:25]
                # Make sure sheetname is uniq
                sheetname_new = self.make_uniq_key(sheetname_new, sheetname_new_list)
                sheetname_new_list.append(sheetname_new)

                print("Copy sheet:", "'%s'"%sheetname_new, cell_from, cell_to, "'%s'"%key)
                self.sheet_groups_dict['sheetname_new'].append(self.sheet_groups_dict[ws.title][key] + [sheetname_new])

                # Group row lists
                group_row_range = range(cell_from[1], cell_to[1]+1)
                #print(group_row_range)

                # Create sheet
                target = self.wb.copy_worksheet(ws)

                # Go to the name of the copied sheet and rename
                target_name = ws.title + " Copy"
                ws_target = self.wb[target_name]
                ws_target.title = sheetname_new

                # Delete rows by setting values to None
                # Loop
                for iRow, cCol in enumerate(ws_target.rows):
                    iRow += 1
                    # Continue if iRow is in current group, and not delete values
                    if iRow in group_row_range:
                        continue
                    for iCol, cCell in enumerate(cCol):
                        iCol += 1
                        # Set value to None
                        cCell.value = None
                        # Set empty font, color
                        cCell.font = Font()
                        cCell.fill = PatternFill()
                        cCell.border = Border()

        # Check
        #print(self.sheet_groups_dict['sheetname_new'])

    def Copy_groups_to_sheets_as_new(self):
        # Get the list of current sheet names
        wb_cur_sheets = self.wb.sheetnames
        # Collect new target sheetnames
        self.sheet_groups_dict['sheetname_new'] = []

        # Loop through worksheets
        for ws in self.wb:
            keys = self.sheet_groups_dict[ws.title]['keys']
            if self.debug_sheet:
                #keys = keys[:1]
                keys = [keys[0]]+[keys[4]]
            for key in keys:
                # Extract
                groupname, cell_from, cell_to = self.sheet_groups_dict[ws.title][key]
                # Define new sheetname
                if groupname in wb_cur_sheets:
                    sheetname_new = ws.title + "_" + groupname
                else:
                    sheetname_new = groupname
                print("Copy sheet:", "'%s'"%sheetname_new, cell_from, cell_to, "'%s'"%key)
                self.sheet_groups_dict['sheetname_new'].append(self.sheet_groups_dict[ws.title][key] + [sheetname_new])

                # Group row lists
                group_row_range = range(cell_from[1], cell_to[1]+1)
                #print(group_row_range)

                # Create new sheet
                ws_target = self.wb.create_sheet(sheetname_new)

                # Loop
                iRow_new = 1
                for iRow, cCol in enumerate(ws.rows):
                    iRow += 1
                    # If row in group range
                    if iRow in group_row_range:
                        for iCol, cCell in enumerate(cCol):
                            iCol += 1
                            # Convert letter
                            iCol_L = get_column_letter(iCol)
                            # Get new cell
                            cell_new = ws_target["%s%s"%(iCol_L, iRow_new)]
                            cell_from = ws["%s%s"%(iCol_L, iRow)]
                            # Copy value
                            cell_new.value = cell_from.value
                            # Copy all styles
                            if cell_from.has_style:
                                cell_new.font = copy.copy(cell_from.font)
                                cell_new.border = copy.copy(cell_from.border)
                                cell_new.fill = copy.copy(cell_from.fill)
                                cell_new.number_format = copy.copy(cell_from.number_format)
                                cell_new.protection = copy.copy(cell_from.protection)
                                cell_new.alignment = copy.copy(cell_from.alignment)
                                #print(dir(cell_from))

                        # Counter
                        iRow_new += 1

        # Check
        #print(self.sheet_groups_dict['sheetname_new'])

    def Find_sheet_groups(self):
        # Loop through worksheets
        sheet_groups_dict = {}
        for ws in self.wb:
            # Make key with Sheet title
            sheet_groups_dict[ws.title] = {}
            sheet_groups_dict_ws_title_keys = []
            sheet_groups_dict_ws_title_keys_uniq = []

            # Get the merged cells of sheet, sorted after columns
            merged_cells = sorted(ws.merged_cells)
            # Split up
            merged_cells_split = []
            for cell_range in merged_cells:
                coord_split = cell_range.coord.split(":")
                #print(ws[coord_split[0]].value)
                merged_cells_split.append(coordinate_from_string(coord_split[0]))
            # Sort after row number
            merged_cells_split.sort(key=lambda x: x[1])

            # Get all groups with (pct)
            freq_groups = []
            for iRow, cCol in enumerate(ws.rows):
                iRow += 1
                cCell = ws["B%s"%(iRow)]
                if cCell.value == None:
                    continue
                # Only if fontsize 8, since the title of groups has this
                if cCell.font.sz == 8:
                    if cCell.value == "(abs)":
                        freq_groups.append((cCell.column, cCell.row))

            # Join together
            freq_merged = freq_groups + merged_cells_split
            # Sort after row number
            freq_merged.sort(key=lambda x: x[1])

            # Make key before looping over rows
            key_cur = None
            key_prev = None
            cCell_val_ascii_prev = None

            # Loop
            for iCol_L, iRow in freq_merged:
                # Get the column number from Column Letter
                iCol = column_index_from_string(iCol_L);
                # Get the Cell object from string method
                cCell = ws["%s%s"%(iCol_L, iRow)]
                #print(cCell, "R:%i C:%i    :"%(iRow, iCol), cCell.value)
                # Get the value
                cCell_val = cCell.value

                # If value of cell is not None
                if cCell_val != None:
                    # Only if Column b
                    if iCol_L == "B":
                        # Only if fontsize 8, since the title of groups has this
                        if cCell.font.sz == 8:
                            # Store current values
                            cCell_val_store = cCell_val
                            iCol_L_store = iCol_L
                            iRow_store = iRow

                            # Convert val to Ascii
                            cCell_val_ascii = cCell_val.encode('ascii', 'ignore').decode("utf-8")
                            if cCell_val_ascii != cCell_val_ascii_prev:
                                # Update
                                cCell_val_ascii_prev = cCell_val_ascii
                                key_cur = self.make_uniq_key(cCell_val_ascii, sheet_groups_dict_ws_title_keys)

                            # Set key first time with list
                            if key_prev == None:
                                key_prev = key_cur

                                # Create empty list and add info
                                sheet_groups_dict_ws_title_keys.append(key_prev)
                                sheet_groups_dict[ws.title][key_prev] = []
                                sheet_groups_dict[ws.title][key_prev].append(cCell_val_store)
                                sheet_groups_dict[ws.title][key_prev].append((iCol_L_store, iRow_store))

                            # If same key, continue
                            if key_cur == key_prev:
                                continue

                            # If keys are different, store to previous and current
                            elif key_cur != key_prev:
                                # Add to previous
                                sheet_groups_dict[ws.title][key_prev].append((iCol_L, iRow-1))
                                # Add cell info
                                sheet_groups_dict_ws_title_keys.append(key_cur)
                                sheet_groups_dict[ws.title][key_cur] = []
                                sheet_groups_dict[ws.title][key_cur].append(cCell_val_store)
                                sheet_groups_dict[ws.title][key_cur].append((iCol_L_store, iRow_store))
                                # Update key
                                key_prev = key_cur

            # After collection per sheet
            # Store last info
            sheet_groups_dict[ws.title][key_prev].append((iCol_L_store, iRow))

            # Add collected keys
            sheet_groups_dict[ws.title]['keys'] = sheet_groups_dict_ws_title_keys
            #print(sheet_groups_dict[ws.title])
            #print(sheet_groups_dict_ws_title_keys)

        # store
        self.sheet_groups_dict = sheet_groups_dict

    def make_uniq_key(self, key=None, keylist=None):
        # If key not in list
        key_in_list = key in keylist
        if not key_in_list:
            return key
        else:
            i = 2
            while key_in_list:
                newkey = key + "_%i"%i
                key_in_list = newkey in keylist
                i += 1
            return newkey

    def Format_Cells_number_format(self, fontsize=None):
        # Loop through worksheets
        for ws in self.wb:
            for iRow, cCol in enumerate(ws.rows):
                iRow += 1
                for iCol, cCell in enumerate(cCol):
                    iCol += 1
                    #print(cCell, "R:%i C:%i    :"%(iRow, iCol), cCell.value)
                    ### Convert format
                    # Get the format
                    cCell_format = cCell.number_format
                    # Convert format
                    if cCell_format == "0.00%":
                        #print(cCell.number_format, type(cCell.number_format))
                        cCell.number_format = "0.0%"

    def Page_Layout_setup(self):
        # Loop through worksheets
        for ws in self.wb:
            # Change from 'portrait' to 'landscape'
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            #print(ws.page_setup.orientation, "\n\n"); [print(i) for i in dir(ws)]
            #print(ws.page_margins) # Narrow: left=0.25, right=0.25, top=0.75, bottom=0.75, header=0.3, footer=0.3
            ws.page_margins.left=0.25; ws.page_margins.right=0.25;
            ws.page_margins.top=0.75; ws.page_margins.bottom=0.75;
            ws.page_margins.header=0.3; ws.page_margins.footer=0.3;

    def delete_empty_sheets(self):
        # Loop through worksheets
        for ws in self.wb:
            #print(ws.title, ws.max_row, ws.max_column)
            # Delete sheet if empty
            if ws.max_row == 1 and ws.max_column == 1:
                #print(ws.title)
                std=self.wb[ws.title]
                self.wb.remove(std)

    def load_wb(self):
        # Open workbook
        self.wb = load_workbook(self.excel_dst)

    def copy_excel(self):
        filename_src, fileext = os.path.splitext(self.excel_src)
        filename_dst = filename_src + "_" +  self.cur_time
        # New destination
        if self.excel_dst == None:
            self.excel_dst = filename_dst+fileext
        # Copy
        shutil.copy2(self.excel_src, self.excel_dst)

