################################################################################
# Copyright (C) Troels Schwarz-Linnet - All Rights Reserved
# Written by Troels Schwarz-Linnet <tlinnet@gmail.com>, January 2018
# 
# Unauthorized copying of this file, via any medium is strictly prohibited.
#
# Any use of this code is strictly unauthorized without the written consent
# by the the author. This code is proprietary of the author.
# 
################################################################################
import shutil, datetime, os, os.path, sys

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Font, PatternFill, Border

# Print version
#from openpyxl import __version__; print(__version__)

from textblob.classifiers import NaiveBayesClassifier


class text:
    def __init__(self, excel_src=None, excel_dst=None):
        # Store 
        self.excel_src = excel_src
        self.excel_dst = excel_dst

        # Make current time
        #self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W-H%H-M%M-S%S")
        self.cur_time = datetime.datetime.now().strftime("%Y-%m-%d-Week%W")

    def run_all(self):
        # Copy Excel workbook
        self.copy_excel()
        # Print all objects of workbook
        #[print(i) for i in dir(self.wb)]
        #print(self.wb.active, self.wb.properties)

        # Open workbook
        self.load_wb()

        # Delete empty sheets
        self.delete_empty_sheets()

        # Read data
        self.read_data()

        # Now train
        self.do_train()

        # Now classify all sentences
        self.do_classify_sentences()

        # Now write to worksheet
        self.write_results()

        # Save Workbook
        self.wb.save(self.excel_dst)


    def write_results(self):
        # Loop over sentences
        sent_irows, sentences = self.sent_dat
        for i, sentence in enumerate(sentences):
            # Extract and unpack
            sent_irow = sent_irows[i]
            iRow, iCol = sent_irow
            # Extract and unpack
            sent_classification = self.sent_classifications[iRow-1]
            classi_sent, classi_class = sent_classification
            # Sanity check
            if sentence != classi_sent:
                print("Oh no!!! Error! %s != %s" %(sentence, classi_sent) )
                continue

            # Get the cell
            cCell = self.ws['%s%i'%(self.index_class_L, iRow+1)]
            # Store the classification
            cCell.value = classi_class

        # Loop over train
        train_irows, train_sentences = self.train_dat
        correct_train = []
        for i, sentence_pair in enumerate(train_sentences):
            sentence, classi_train = sentence_pair
            # Extract and unpack
            train_irow = train_irows[i]
            iRow, iCol_1, iCol_2 = train_irow
            # Extract and unpack
            sent_classification = self.sent_classifications[iRow-1]
            classi_sent, classi_class = sent_classification
            # Sanity check
            if sentence != classi_sent:
                print("Oh no!!! Error! %s != %s" %(sentence, classi_sent) )
                continue
            # Check
            if classi_train == classi_class:
                equal = 1
            else:
                equal = 0
            # Add
            correct_train.append(equal)
            # Get the cell
            cCell = self.ws['%s%i'%(self.index_correct_train_L, iRow+1)]
            # Store the classification
            cCell.value = equal

        # Write results
        self.ws['J1'].value = "correct_train"
        self.ws['J2'].value = "sum"
        self.ws['K2'].value = sum(correct_train)
        self.ws['J3'].value = "total"
        self.ws['K3'].value = len(correct_train)
        self.ws['J4'].value = "pct"
        self.ws['K4'].value = float("%0.2f"%( sum(correct_train)/len(correct_train) ))

        # Loop over test
        correct_test = []
        if self.has_test:
            test_irows, test_sentences = self.test_dat
            for i, sentence_pair in enumerate(test_sentences):
                sentence, classi_test = sentence_pair
                # Extract and unpack
                test_irow = test_irows[i]
                iRow, iCol_1, iCol_2 = test_irow
                # Extract and unpack
                sent_classification = self.sent_classifications[iRow-1]
                classi_sent, classi_class = sent_classification
                # Sanity check
                if sentence != classi_sent:
                    print("Oh no!!! Error! %s != %s" %(sentence, classi_sent) )
                    continue
                # Check
                if classi_test == classi_class:
                    equal = 1
                else:
                    equal = 0
                # Add
                correct_test.append(equal)
                # Get the cell
                cCell = self.ws['%s%i'%(self.index_correct_test_L, iRow+1)]
                # Store the classification
                cCell.value = equal

            # Write results
            self.ws['J7'].value = "correct_test"
            self.ws['J8'].value = "sum"
            self.ws['K8'].value = sum(correct_test)
            self.ws['J9'].value = "total"
            self.ws['K9'].value = len(correct_test)
            self.ws['J10'].value = "pct"
            self.ws['K10'].value = float("%0.2f"%( sum(correct_test)/len(correct_test) ))


    def do_classify_sentences(self):
        sentences = self.sent_dat[1]
        
        # Loop over all sentences
        self.sent_classifications = []
        for sent in sentences:
            # Do classification
            classification = self.cl.classify(sent)
            # Store
            self.sent_classifications.append((sent, classification))

    def do_train(self):
        # Get the data
        train = self.train_dat[1]

        # http://textblob.readthedocs.io/en/dev/classifiers.html#classifying-text
        # Now we’ll create a Naive Bayes classifier, passing the training data into the constructor.
        self.cl = NaiveBayesClassifier(train)

        # Score on own train
        self.accuracy_train = self.cl.accuracy(train)
        print("After training %i sentences, the accuracy on %i train sentences is %0.2f"% (len(train), len(train), self.accuracy_train) )

        # If we have test data, we can score the classifier
        if self.has_test:
            test = self.test_dat[1]
            self.accuracy_test = self.cl.accuracy(test)
            print("After training %i sentences, the accuracy on %i test sentences is %0.2f"% (len(train), len(test), self.accuracy_test) )

    def read_data(self):
        # Use the first sheet
        ws_name = self.wb.sheetnames[0]
        self.ws = self.wb[ws_name]
        
        # Get row 1
        self.header_vals = []
        for iCol in range(1, self.ws.max_column):
            iCol_L = get_column_letter(iCol)
            cCell = self.ws['%s1'%(iCol_L)]
            self.header_vals.append(cCell.value)
        # Check for header
        if 'train' in self.header_vals:
            self.has_header = True
            self.iRow_skip = 1
        else:
            self.has_header = False
            self.header_vals = ['sentences', 'train', 'classify', 'test', 'correct_train', 'correct_test']
            self.iRow_skip = 0

        # Get the index of columns
        if 'sentences' in self.header_vals:
            self.index_sent = self.header_vals.index('sentences')
        else:
            self.index_sent = 0

        if 'train' in self.header_vals:
            self.index_train = self.header_vals.index('train')
        else:
            self.index_train = 1

        if 'classify' in self.header_vals:
            self.index_class = self.header_vals.index('classify')
        else:
            self.index_class = 2
        self.index_class_L = get_column_letter(self.index_class+1)

        if 'test' in self.header_vals:
            self.index_test = self.header_vals.index('test')
            self.has_test = True
        else:
            self.index_test = 3
            self.has_test = False

        if 'correct_train' in self.header_vals:
            self.index_correct_train = self.header_vals.index('correct_train')
        else:
            self.index_correct_train
        self.index_correct_train_L = get_column_letter(self.index_correct_train+1)

        if 'correct_test' in self.header_vals:
            self.index_correct_test = self.header_vals.index('correct_test')
        else:
            self.index_correct_test
        self.index_correct_test_L = get_column_letter(self.index_correct_test+1)

        # Collect all sentences
        self.sent_dat = self.collect_column_data(self.index_sent)

        # Collect paired values of sentences and train
        self.train_dat = self.collect_column_data(self.index_sent, self.index_train)

        # If test is available
        if self.has_test:
            self.test_dat = self.collect_column_data(self.index_sent, self.index_test)


    def collect_column_data(self, col_i_1=1, col_i_2=None):
        # Loop over rows
        iRow_nrs = []
        iRow_pair_vals = []
        
        for iRow, cCol in enumerate(self.ws.rows):
            # Possible skip
            if iRow < self.iRow_skip:
                continue
            if col_i_2 != None:
                # Get the cells
                cCell_1 = cCol[col_i_1]
                cCell_2 = cCol[col_i_2]
                # Get the value
                val_1 = cCell_1.value
                val_2 = cCell_2.value
                if val_1 != None and val_2 != None:
                    iRow_nrs.append((iRow,col_i_1,col_i_2))
                    iRow_pair_vals.append((val_1, val_2))

            elif col_i_2 == None:
                # Get the cells
                cCell_1 = cCol[col_i_1]
                # Get the value
                val_1 = cCell_1.value
                if val_1 != None:
                    iRow_nrs.append((iRow,col_i_1))
                    iRow_pair_vals.append(val_1)

        if len(iRow_nrs) > 0:
            return [iRow_nrs, iRow_pair_vals]
        else:
            return None

    def delete_empty_sheets(self):
        # Loop through worksheets
        for ws in self.wb:
            #print(ws.title, ws.max_row, ws.max_column)
            # Delete sheet if empty
            if ws.max_row == 1 and ws.max_column == 1:
                #print(ws.title)
                std=self.wb[ws.title]
                self.wb.remove(std)

    def load_wb(self):
        # Open workbook
        self.wb = load_workbook(self.excel_dst, data_only=True)

    def copy_excel(self):
        filename_src, fileext = os.path.splitext(self.excel_src)
        filename_dst = filename_src + "_" +  self.cur_time
        # New destination
        if self.excel_dst == None:
            # If this is an update on file with date in it
            if self.cur_time in filename_src:
                filename_split = filename_src.split(self.cur_time)
                if filename_split[-1] == "":
                    filename_dst = filename_split[0] + self.cur_time + "_002"
                else:
                    version = int(filename_split[-1].split("_")[-1])
                    version += 1
                    filename_dst = filename_split[0] + self.cur_time + "_%03d"%(version)
            
            # Create the file
            self.excel_dst = filename_dst+fileext

        # Copy, but do not overwrite !
        if os.path.isfile(self.excel_dst):
            print("File %s does already exists! Exiting!"%(self.excel_dst))
            sys.exit()
        else:
            shutil.copy2(self.excel_src, self.excel_dst)
