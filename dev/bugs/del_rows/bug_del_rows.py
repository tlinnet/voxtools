import shutil, sys
from distutils.version import StrictVersion

from openpyxl import Workbook
from openpyxl import load_workbook

# Test version
from openpyxl import __version__
test_version = StrictVersion(__version__) == StrictVersion("2.5.0")
if not test_version:
    print("You need to have openpyxl version 2.5.0. You have %s"%__version__)
    sys.exit()

# Create Workbook, Get the active shees
wb = Workbook()
ws = wb.active

ws.append(['Hello 1'])
ws.append(['Rem 1'])
ws.append(['Rem 2'])
ws.append(['Rem 3'])
ws.append(['Hello 2'])

# Modify some properties
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 30
ws.row_dimensions[3].height = 40
ws.row_dimensions[4].height = 50
ws.row_dimensions[5].height = 60

# Save & Copy
wb.save('test_1.xlsx')
shutil.copy2('test_1.xlsx', 'test_2.xlsx')

# Open workbook
wb = load_workbook('test_2.xlsx')
ws = wb.active

# Check for Rem and set height
for iRow, cCol in enumerate(ws.rows):
    iRow += 1
    cCell = ws['A%i'%(iRow)]
    if cCell.value == None:
        continue
    elif "Rem" in cCell.value:
        # Set height property we can look for
        cCell.value = "DELETE ME"

# Loop again, and look for height
cont_loop = True
while cont_loop:
    for iRow, cCol in enumerate(ws.rows):
        iRow += 1
        cCell = ws['A%i'%(iRow)]

        # Break at end
        if iRow == ws.max_row:
            cont_loop = False
        elif cCell.value == "DELETE ME":
            # Delete
            ws.delete_rows(iRow, 1)
            break

wb.save('test_2.xlsx')